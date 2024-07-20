from institute.models import Block
from students.models import Attendance
from django.utils import timezone
# قراءة وكتابة ملفات Excel (كلا التنسيقات .xlsx و .xlsm)
# يتيح لك العمل مع ملفات Excel برمجيًا ، وإنشاء جداول بيانات جديدة ، وتعديل الملفات الموجودة ، وتطبيق أنماط وتنسيقات مختلفة.
#   workbook   عبارة عن مجموعة من الأوراق بشكل أساسي
from openpyxl import Workbook
# styles  تحتوي وحدة الأنماط على فئات ووظائف متعلقة بأنماط الخلايا وتنسيقها
from openpyxl import styles





# إنشاء مصنف
class AttendanceBookGenerator:
    def __init__(self, block_id, year_month):
        self.block_id = block_id
        self.year_month = year_month
    
    def generate_workbook(self):
        workbook  = Workbook()
        # يزيل هذا السطر الورقة النشطة الافتراضية من المصنف.ويترك المصنف بدون أي أوراق
        workbook.remove(workbook.active)

        if self.block_id == 'all':
            for block in Block.objects.all():
                # من اجل كل كتله يهيئ غرض من الصنف الذي ينشئ ورقة الحضور ويقسم التاريخ الى شهر وسنه
                SheetGenerator = BlockAttendanceSheetGenerator(block.id, self.year_month)
                # الان ضخ المعلومات في ورقة الاكسيل
                SheetGenerator.generate_block_sheet(workbook)
        else:
            SheetGenerator = BlockAttendanceSheetGenerator(self.block_id, self.year_month)
            SheetGenerator.generate_block_sheet(workbook)

        return workbook

# إنشاء أوراق الحضور لكل كتلة
class BlockAttendanceSheetGenerator:
    def __init__(self, block_id, year_month):
        self.block = Block.objects.get(id = block_id)
        # يقوم بجلب اغراض االاتيندانس الموافقة للكتلة والتي توافق ان يكون رقم الطالب في الاتيندانس موافق لرقم الطالب في الغرفة في البلوك
        self.attendance_list = Attendance.objects.filter(student__in = self.block.roomdetail_set.all().values_list('student', flat=True))
        # print(str(self.attendance_list) + self.block.name)  # <QuerySet [<Attendance: 222222>, <Attendance: 333333>]>block6
        if year_month == 'all':
            self.month = 'all'
            self.year = 'all'
        else:
            # [2023 ، 8]
            self.year, self.month = [int(x) for x in year_month.split("-")]

    def generate_block_sheet(self, workbook):
        # انشاء ورقة عمل جديدة
        # يتطلب عدة وسيطات اختيارية ، مثل العنوان والفهرس و header_rows. هنا ، نستخدم وسيطة العنوان لتحديد اسم ورقة العمل ، والتي تم تعيينها على اسم الكتلة.
        worksheet = workbook.create_sheet(title = "{}".format(self.block.name))

        # مسؤولة عن حساب قائمة التواريخ التي سيتم استخدامها كرؤوس أعمدة في ورقة الحضور
        self.generate_dates()       #[datetime.datetime(2023, 8, 1, 0, 0),....., datetime.datetime(2023, 8, 31, 0, 0)]
        # رؤوس مُنشأة ديناميكيًا لكل تاريخ
        headers = ['Regd. No.', 'Name'] + [date.strftime("%d/%m/%y") for date in self.generated_dates]
        # يقوم هذا السطر بتهيئة متغير row_num إلى 1. سيتم استخدامه لتتبع رقم الصف الحالي في ورقة الحضور أثناء ملء البيانات
        row_num = 1

        #  يتكرر هذا for loop فوق قائمة الرؤوس ، ولكل عنصر (عنوان عمود) في القائمة ، فإنه يعين الفهرس الحالي (المستند إلى 1) إلى col_num وقيمة عنوان العمود إلى column_title.
        for col_num, column_title in enumerate(headers, 1):
            #  يحصل هذا السطر علي الخلية الموجودة في ورقه العمل (ورقه Excel) في row_num و col_num المحدد.
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.font = styles.Font(bold = True)
            cell.value = column_title

        for attendance in self.attendance_list:
            row_num += 1
            # get_student_attendance يقوم بامرار حلقة على اعمده التاريخ في ورقة الاكسل ومن اجل كل تاريخ يقوم بتحويله الى الصيغة المناسبة
            # ثم يبحث في كل من سجلات الغياب وسجلات الحضور عن هذا التاريخ فاذا وجده يضع حرف بي او اااا او شحطه
            # وبنهاية الحلقة تكون العلامات والتواريخ ضمن الليست بريسينت ابسينت ليست
            row_data = [attendance.student.regd_no, attendance.student.name] + self.get_student_attendance(attendance)

            for col_num, cell_value in enumerate(row_data, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value
                if cell_value == 'P': cell.font = styles.Font(color='28A745')
                elif cell_value == 'A': cell.font = styles.Font(bold=True, color='DC3545')


    def get_student_attendance(self, attendance):
        present_absent_list = []
        for date in self.generated_dates:
            date_formatted = date.strftime("%Y %m %d")
            if attendance.present_dates and attendance.present_dates.find(date_formatted) != -1:
                present_absent_list.append('P')
            elif attendance.absent_dates and attendance.absent_dates.find(date_formatted) != -1:
                present_absent_list.append('A')
            else:
                present_absent_list.append('-')

        return present_absent_list

    def generate_dates(self):
        if self.year == 'all' or self.month == 'all':
            # ستعيد مجموعه من اجل كل غرفة احادية ثنائية رباعية تحوي تواريخ الحضور والغياب لكل غرفة
            date_set = self.get_marked_dates()
            # print('date_set' + str(date_set))               #date_set{datetime.datetime(2023, 8, 2, 0, 0), datetime.datetime(2023, 8, 1, 0, 0), datetime.datetime(2023, 8, 5, 0, 0)}
            # هنا تعمل الدالة على وضع متغيرين سنه وشهر
            month_set = self.get_month_year_set(date_set)
            # print('month_set' + str(month_set))             #month_set{(8, 2023)}
            generated_dates = []

            for item in month_set:
                # توليد قائمه بايام الشهر والسنه الممررين
                dates_of_month = self.get_dates_of_month(item[0], item[1])
                # print('dates_of_month' + str(dates_of_month))               #[datetime.datetime(2023, 8, 1, 0, 0),......, datetime.datetime(2023, 8, 31, 0, 0)]
                generated_dates += dates_of_month

            self.generated_dates = sorted(generated_dates)
            # print("self.generated_dates" + str(generated_dates))
        
        else:
            self.generated_dates = self.get_dates_of_month(self.month, self.year)

    # الغرض من هذه الطريقة هو إنشاء قائمة بجميع التواريخ خلال الشهر والسنة المحددين. إنه مفيد بشكل خاص لإنشاء رؤوس في ورقة الحضور ، حيث يتوافق كل عمود مع تاريخ في الشهر. تتيح لك الطريقة إنشاء جميع التواريخ في الشهر بحيث يمكنك ملء بيانات الحضور لكل طالب في كل تاريخ في ورقة
    def get_dates_of_month(self, month, year):
            # اليوم: يتم تعيين هذا المتغير على كائن timedelta الذي يمثل يومًا واحدًا. يتم استخدامه لاحقًا لزيادة التاريخ.
            day = timezone.timedelta(days=1)
            # يحتفظ هذا المتغير باليوم الأول من الشهر والسنة المحددين
            # يرجع عنصر تاريخ ووقت بالسنة والشهر واليوم المحددين
            start_date = timezone.datetime(year = year, month = month, day = 1)
            #  ستخزن هذه القائمة جميع التواريخ خلال الشهر المحدد
            dates_of_month = []
            d = start_date
            while d.month == month:
                dates_of_month.append(d)
                d += day

            return sorted(dates_of_month)

    def get_marked_dates(self):
        date_set = set()
        for attendance in self.attendance_list:
            if attendance.present_dates:
                date_set |= set(attendance.present_dates.split(','))
            if attendance.absent_dates:
                date_set |= set(attendance.absent_dates.split(','))
        date_set.discard('')

        date_format = "%Y %m %d"
        # هذا السطر بفضل الله ضفتو لان مكان متساوي التنسيق بين التنين
        date_set = set([date.strip() for date in date_set])  # Remove leading/trailing spaces from date strings
        date_set = set([timezone.datetime.strptime(date, date_format) for date in date_set])
        self.marked_dates = date_set

        return self.marked_dates

    # Set of tuples (month, year) that have attendance marked
    def get_month_year_set(self, date_set):
        month_year_set = set([(date.month, date.year) for date in date_set])
        return month_year_set